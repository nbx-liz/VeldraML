"""Service helpers used by the Dash adapter."""

from __future__ import annotations

import difflib
import json
import logging
import os
import shutil
import threading
from contextlib import contextmanager
from dataclasses import asdict, dataclass, is_dataclass
from datetime import UTC, datetime, timedelta, timezone
from pathlib import Path
from time import perf_counter
from typing import Any, Literal

import pandas as pd
import yaml
from pydantic import ValidationError

from veldra.api.exceptions import (
    VeldraArtifactError,
    VeldraNotImplementedError,
    VeldraValidationError,
)
from veldra.api.logging import log_event
from veldra.config.migrate import migrate_run_config_file, migrate_run_config_payload
from veldra.config.models import RunConfig
from veldra.gui._lazy_runtime import (
    resolve_artifact_class,
    resolve_data_loader,
    resolve_runner_function,
)
from veldra.gui.job_store import GuiJobStore
from veldra.gui.types import (
    ArtifactSpec,
    ArtifactSummary,
    GuiJobPriority,
    GuiJobRecord,
    GuiJobResult,
    GuiRunResult,
    PaginatedResult,
    RetryPolicy,
    RunInvocation,
)

LOGGER = logging.getLogger("veldra.gui.services")
_RUNTIME_LOCK = threading.Lock()
_JOB_STORE: GuiJobStore | None = None
_JOB_WORKER: Any | None = None
_JST = timezone(timedelta(hours=9))

# Lazy runtime symbols for heavyweight deps.
_ARTIFACT_CLS: Any | None = None
evaluate: Any | None = None
estimate_dr: Any | None = None
export: Any | None = None
fit: Any | None = None
predict: Any | None = None
simulate: Any | None = None
tune: Any | None = None
load_tabular_data: Any | None = None


def _get_artifact_cls() -> Any:
    global Artifact, _ARTIFACT_CLS
    if Artifact is not _ArtifactProxy:
        return Artifact
    _ARTIFACT_CLS = resolve_artifact_class(cached_class=_ARTIFACT_CLS)
    return _ARTIFACT_CLS


class _ArtifactProxy:
    @staticmethod
    def load(path: str) -> Any:
        return _get_artifact_cls().load(path)


# Backward-compat name used in tests via monkeypatch("...Artifact.load", ...).
Artifact: Any = _ArtifactProxy


def _get_runner_func(name: str) -> Any:
    global evaluate, estimate_dr, export, fit, predict, simulate, tune
    current = {
        "evaluate": evaluate,
        "estimate_dr": estimate_dr,
        "export": export,
        "fit": fit,
        "predict": predict,
        "simulate": simulate,
        "tune": tune,
    }.get(name)
    if current is not None:
        return current
    resolved = resolve_runner_function(name)
    if name == "evaluate":
        evaluate = resolved
    elif name == "estimate_dr":
        estimate_dr = resolved
    elif name == "export":
        export = resolved
    elif name == "fit":
        fit = resolved
    elif name == "predict":
        predict = resolved
    elif name == "simulate":
        simulate = resolved
    elif name == "tune":
        tune = resolved
    return resolved


def _get_load_tabular_data() -> Any:
    global load_tabular_data
    load_tabular_data = resolve_data_loader(current_loader=load_tabular_data)
    return load_tabular_data


def inspect_data(path: str) -> dict[str, Any]:
    """Inspect a data file and return preview and statistics."""
    try:
        data_path = Path(_require(path, "path"))
        if not data_path.exists():
            raise VeldraValidationError(f"Data file does not exist: {data_path}")
        loader = _get_load_tabular_data()
        df = loader(str(data_path))

        # Calculate column stats
        numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
        cat_cols = df.select_dtypes(exclude=["number"]).columns.tolist()
        datetime_cols = df.select_dtypes(include=["datetime", "datetimetz"]).columns.tolist()

        column_profiles: list[dict[str, Any]] = []
        warnings: list[str] = []
        for col in df.columns:
            series = df[col]
            missing_rate = float(series.isna().mean())
            unique_count = int(series.nunique(dropna=True))
            dtype_name = str(series.dtype)
            inferred_kind = (
                "datetime"
                if col in datetime_cols
                else ("numeric" if col in numeric_cols else "categorical")
            )
            col_info = {
                "name": str(col),
                "dtype": dtype_name,
                "kind": inferred_kind,
                "missing_rate": missing_rate,
                "unique_count": unique_count,
                "constant": unique_count <= 1,
                "high_missing": missing_rate > 0.5,
                "high_cardinality": inferred_kind == "categorical" and unique_count > 100,
            }
            column_profiles.append(col_info)
            if col_info["high_missing"]:
                warnings.append(f"High missing rate: {col} ({missing_rate:.1%})")
            if col_info["constant"]:
                warnings.append(f"Constant column: {col}")
            if col_info["high_cardinality"]:
                warnings.append(f"High cardinality: {col} ({unique_count})")

        stats = {
            "n_rows": len(df),
            "n_cols": len(df.columns),
            "columns": list(df.columns),
            "numeric_cols": numeric_cols,
            "categorical_cols": cat_cols,
            "datetime_cols": datetime_cols,
            "missing_count": int(df.isnull().sum().sum()),
            "memory_usage_mb": float(df.memory_usage(deep=True).sum() / 1024 / 1024),
            "column_profiles": column_profiles,
            "warnings": warnings,
        }

        return {
            "success": True,
            "stats": stats,
            "preview": [],
            "path": str(data_path.resolve()),
        }
    except Exception as exc:
        return {
            "success": False,
            "error": str(exc),
        }


def record_perf_metric(name: str, elapsed_ms: float, payload: dict[str, Any] | None = None) -> None:
    data = {
        "metric": str(name),
        "elapsed_ms": round(float(elapsed_ms), 3),
    }
    if payload:
        data.update(payload)
    if float(elapsed_ms) >= 100.0:
        LOGGER.warning("gui_perf_metric_slow", extra=data)
    else:
        LOGGER.info("gui_perf_metric", extra=data)


def load_data_preview_page(
    path: str,
    *,
    offset: int = 0,
    limit: int = 200,
    columns: list[str] | None = None,
) -> PaginatedResult[dict[str, Any]]:
    started_at = perf_counter()
    data_path = Path(_require(path, "path"))
    if not data_path.exists():
        raise VeldraValidationError(f"Data file does not exist: {data_path}")
    loader = _get_load_tabular_data()
    df = loader(str(data_path))
    safe_offset = max(0, int(offset))
    safe_limit = max(1, min(int(limit), 2000))
    selected_df = df
    if columns:
        keep = [str(col) for col in columns if str(col) in df.columns]
        if keep:
            selected_df = df.loc[:, keep]
    sliced = selected_df.iloc[safe_offset : safe_offset + safe_limit]
    payload = (
        sliced.astype(object).where(pd.notnull(sliced), None).to_dict(orient="records")
        if not sliced.empty
        else []
    )
    result = PaginatedResult[dict[str, Any]](
        items=payload,
        total_count=int(len(selected_df)),
        limit=safe_limit,
        offset=safe_offset,
    )
    record_perf_metric(
        "load_data_preview_page",
        (perf_counter() - started_at) * 1000.0,
        {
            "path": str(data_path),
            "offset": safe_offset,
            "limit": safe_limit,
            "total_count": result.total_count,
        },
    )
    return result


def default_job_db_path() -> Path:
    return Path(os.getenv("VELDRA_GUI_JOB_DB_PATH", ".veldra_gui/jobs.sqlite3"))


def set_gui_runtime(*, job_store: GuiJobStore, worker: Any | None) -> None:
    global _JOB_STORE, _JOB_WORKER
    with _RUNTIME_LOCK:
        _JOB_STORE = job_store
        _JOB_WORKER = worker


def get_gui_job_store() -> GuiJobStore:
    global _JOB_STORE
    with _RUNTIME_LOCK:
        if _JOB_STORE is None:
            _JOB_STORE = GuiJobStore(default_job_db_path())
        return _JOB_STORE


def stop_gui_runtime() -> None:
    global _JOB_STORE, _JOB_WORKER
    with _RUNTIME_LOCK:
        worker = _JOB_WORKER
        _JOB_WORKER = None
        _JOB_STORE = None
    if worker is not None and hasattr(worker, "stop"):
        worker.stop()


def run_housekeeping_cycle(
    *,
    archive_ttl_days: int = 30,
    purge_ttl_days: int = 90,
    batch_size: int = 200,
) -> dict[str, int]:
    now = datetime.now(UTC)
    archive_cutoff = (now - timedelta(days=max(1, int(archive_ttl_days)))).isoformat()
    purge_cutoff = (now - timedelta(days=max(1, int(purge_ttl_days)))).isoformat()
    store = get_gui_job_store()
    moved = store.archive_jobs(cutoff_utc=archive_cutoff, batch_size=batch_size)
    purged = store.purge_archived_jobs(cutoff_utc=purge_cutoff, batch_size=batch_size)
    return {"archived_jobs": int(moved), "purged_archived_jobs": int(purged)}


def _start_worker_if_needed() -> None:
    with _RUNTIME_LOCK:
        worker = _JOB_WORKER
    if worker is not None and hasattr(worker, "start"):
        worker.start()


def normalize_gui_error(exc: Exception) -> str:
    if isinstance(exc, VeldraValidationError):
        return f"Validation error: {exc}"
    if isinstance(exc, VeldraArtifactError):
        return f"Artifact error: {exc}"
    if isinstance(exc, VeldraNotImplementedError):
        return f"Not implemented: {exc}"
    return f"{exc.__class__.__name__}: {exc}"


def classify_gui_error(exc: Exception) -> str:
    if isinstance(exc, CanceledByUser):
        return "cancel"
    if isinstance(exc, VeldraValidationError):
        return "validation"
    if isinstance(exc, FileNotFoundError):
        return "file_not_found"
    if isinstance(exc, PermissionError):
        return "permission"
    if isinstance(exc, TimeoutError):
        return "timeout"
    if isinstance(exc, MemoryError):
        return "memory"
    msg = str(exc).lower()
    if "not found" in msg or "does not exist" in msg:
        return "file_not_found"
    if "permission" in msg or "access denied" in msg:
        return "permission"
    if "timeout" in msg or "timed out" in msg:
        return "timeout"
    if "busy" in msg or "database is locked" in msg:
        return "resource_busy"
    if "temporar" in msg or "connection reset" in msg:
        return "io_transient"
    return "unknown"


def build_next_steps(error_kind: str, *, action: str | None = None) -> list[str]:
    if error_kind == "validation":
        return ["設定値と入力パスを確認し、RunConfig を再検証してください。"]
    if error_kind == "file_not_found":
        return ["対象ファイルの存在と相対/絶対パスを確認してください。"]
    if error_kind == "permission":
        return ["ファイル権限と書き込み先ディレクトリ権限を確認してください。"]
    if error_kind == "memory":
        return ["データ量や `num_boost_round` を下げるか、分割数を減らして再実行してください。"]
    if error_kind == "timeout":
        return ["環境負荷を下げて再実行し、必要ならリトライ回数を増やしてください。"]
    if error_kind == "cancel":
        return ["処理はユーザーキャンセルで停止しました。必要なら Retry Task を実行してください。"]
    if error_kind in {"resource_busy", "io_transient"}:
        return ["一時的障害の可能性があります。少し待ってから再実行してください。"]
    if action:
        return [f"ジョブログを確認し、`{action}` の入力と依存関係を再確認してください。"]
    return []


def _build_error_payload(exc: Exception, *, action: str | None = None) -> dict[str, Any]:
    kind = classify_gui_error(exc)
    return {
        "error_kind": kind,
        "next_steps": build_next_steps(kind, action=action),
    }


class CanceledByUser(RuntimeError):
    """Raised when cooperative cancellation is requested for a running GUI job."""


def _invoke_runner_with_optional_hook(
    fn: Any,
    *args: Any,
    cancellation_hook: Any,
) -> Any:
    try:
        return fn(*args, cancellation_hook=cancellation_hook)
    except TypeError as exc:
        if "unexpected keyword argument 'cancellation_hook'" in str(exc):
            return fn(*args)
        raise


def _require(value: str | None, field_name: str) -> str:
    if value is None or not value.strip():
        raise VeldraValidationError(f"{field_name} is required.")
    return value.strip()


def _load_config_from_yaml(yaml_text: str) -> RunConfig:
    raw = yaml.safe_load(yaml_text)
    if not isinstance(raw, dict):
        raise VeldraValidationError("Config YAML must deserialize to an object.")
    return RunConfig.model_validate(raw)


def validate_config(yaml_text: str) -> RunConfig:
    return _load_config_from_yaml(yaml_text)


def validate_config_with_guidance(yaml_text: str) -> dict[str, Any]:
    timestamp = datetime.now(UTC).isoformat()
    try:
        validate_config(yaml_text)
        return {
            "ok": True,
            "errors": [],
            "warnings": [],
            "timestamp_utc": timestamp,
        }
    except ValidationError as exc:
        errors: list[dict[str, Any]] = []
        for item in exc.errors():
            loc = item.get("loc", ())
            path = ".".join(str(seg) for seg in loc if seg is not None) or "root"
            msg = str(item.get("msg") or "Invalid value.")
            errors.append(
                {
                    "path": path,
                    "message": msg,
                    "suggestions": ["設定値と型、必須キーを確認してください。"],
                }
            )
        return {
            "ok": False,
            "errors": errors,
            "warnings": [],
            "timestamp_utc": timestamp,
        }
    except Exception as exc:
        kind = classify_gui_error(exc)
        return {
            "ok": False,
            "errors": [
                {
                    "path": "root",
                    "message": normalize_gui_error(exc),
                    "suggestions": build_next_steps(kind),
                }
            ],
            "warnings": [],
            "timestamp_utc": timestamp,
        }


def load_config_yaml(path: str) -> str:
    config_path = Path(_require(path, "config_path"))
    if not config_path.exists():
        raise VeldraValidationError(f"Config file does not exist: {config_path}")
    return config_path.read_text(encoding="utf-8")


def save_config_yaml(path: str, yaml_text: str) -> str:
    config_path = Path(_require(path, "config_path"))
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(yaml_text, encoding="utf-8")
    return str(config_path)


def migrate_config_from_yaml(
    yaml_text: str,
    target_version: int = 1,
) -> tuple[str, str]:
    """Migrate config from YAML string. Returns (normalized_yaml, diff)."""
    try:
        payload = yaml.safe_load(yaml_text or "")
        if not isinstance(payload, dict):
            raise VeldraValidationError("Config YAML must deserialize to an object.")
        normalized, _migration = migrate_run_config_payload(payload, target_version=target_version)
        normalized_yaml = yaml.safe_dump(normalized, sort_keys=False, allow_unicode=True)

        # Calculate diff
        input_lines = (yaml_text or "").splitlines()
        output_lines = normalized_yaml.splitlines()
        diff = "\n".join(
            difflib.unified_diff(
                input_lines,
                output_lines,
                fromfile="Original",
                tofile="Normalized",
                lineterm="",
            )
        )
        return normalized_yaml, diff
    except Exception as exc:
        raise VeldraValidationError(f"Migration failed: {exc}")


def migrate_config_file_via_gui(
    input_path: str,
    output_path: str | None = None,
    target_version: int = 1,
) -> str:
    """Migrate config file. Returns summary message."""
    try:
        result = migrate_run_config_file(
            input_path=input_path,
            output_path=output_path,
            target_version=target_version,
        )
        return (
            f"Migration successful.\\n"
            f"Source: {result.input_path}\\n"
            f"Output: {result.output_path}\\n"
            f"Version: {result.source_version} -> {result.target_version}\\n"
            f"Changed: {result.changed}"
        )
    except Exception as exc:
        raise VeldraValidationError(f"File migration failed: {exc}")


def _result_to_payload(result: Any) -> dict[str, Any]:
    if is_dataclass(result):
        # Handle EvalResult or others that might have DataFrame in fields
        # asdict fails on DataFrame. We need to handle it manually or use asdict safely
        try:
            payload = asdict(result)
        except Exception:
            # Fallback for when asdict fails (e.g. DataFrame in field)
            payload = {}
            for field in result.__dataclass_fields__:
                val = getattr(result, field)
                if isinstance(val, pd.DataFrame):
                    payload[field] = {
                        "n_rows": int(len(val)),
                        "columns": list(val.columns),
                        "preview": val.head(20).to_dict(orient="records"),
                    }
                else:
                    payload[field] = val
    elif isinstance(result, pd.DataFrame):
        payload = {
            "n_rows": int(len(result)),
            "columns": list(result.columns),
            "preview": result.head(20).to_dict(orient="records"),
        }
    elif isinstance(result, (dict, list, str, int, float, bool)) or result is None:
        payload = {"result": result}
    else:
        payload = {"result_repr": repr(result)}

    data_obj = payload.get("data")
    if isinstance(data_obj, pd.DataFrame):
        payload["data"] = {
            "n_rows": int(len(data_obj)),
            "columns": list(data_obj.columns),
            "preview": data_obj.head(20).to_dict(orient="records"),
        }
    return payload


def _resolve_config(invocation: RunInvocation) -> RunConfig:
    if invocation.config_yaml and invocation.config_yaml.strip():
        return _load_config_from_yaml(invocation.config_yaml)
    if invocation.config_path and invocation.config_path.strip():
        config_text = load_config_yaml(invocation.config_path)
        return _load_config_from_yaml(config_text)
    raise VeldraValidationError("Either config YAML or config path is required.")


def _load_scenarios(path: str) -> Any:
    scenarios_path = Path(_require(path, "scenarios_path"))
    if not scenarios_path.exists():
        raise VeldraValidationError(f"Scenarios file does not exist: {scenarios_path}")
    text = scenarios_path.read_text(encoding="utf-8")
    if scenarios_path.suffix.lower() == ".json":
        return json.loads(text)
    return yaml.safe_load(text)


def _normalize_dtype_bucket(dtype_name: str) -> str:
    text = str(dtype_name).strip().lower()
    if not text:
        return "unknown"
    if "datetime" in text or "date" in text:
        return "datetime"
    if "int" in text or "float" in text or "double" in text or "decimal" in text:
        return "numeric"
    if "bool" in text:
        return "bool"
    if "category" in text or "string" in text or "object" in text:
        return "categorical"
    return "unknown"


def _studio_predict_tmp_dir() -> Path:
    tmp_dir = Path(".veldra_gui") / "tmp"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    return tmp_dir


def infer_task_type(data: pd.DataFrame, target_col: str) -> str:
    """Infer recommended task type from target column characteristics."""
    if target_col not in data.columns:
        return "regression"
    series = data[target_col].dropna()
    if series.empty:
        return "regression"
    n_unique = int(series.nunique())
    if n_unique == 2:
        return "binary"
    is_int_like = pd.api.types.is_integer_dtype(series)
    if 3 <= n_unique <= 20 and is_int_like:
        return "multiclass"
    return "regression"


def get_artifact_spec(artifact_path: str) -> ArtifactSpec:
    path = Path(_require(artifact_path, "artifact_path"))
    if not path.exists() or not path.is_dir():
        raise VeldraValidationError(f"Artifact directory does not exist: {path}")

    manifest_path = path / "manifest.json"
    run_config_path = path / "run_config.yaml"
    feature_schema_path = path / "feature_schema.json"
    metrics_path = path / "metrics.json"
    required = [manifest_path, run_config_path, feature_schema_path]
    missing = [item.name for item in required if not item.exists()]
    if missing:
        raise VeldraValidationError(
            f"Artifact is missing required file(s): {', '.join(missing)}"
        )

    try:
        manifest_obj = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise VeldraValidationError(f"Invalid manifest.json: {exc}") from exc
    try:
        config_obj = yaml.safe_load(run_config_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise VeldraValidationError(f"Invalid run_config.yaml: {exc}") from exc
    try:
        feature_schema_obj = json.loads(feature_schema_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise VeldraValidationError(f"Invalid feature_schema.json: {exc}") from exc

    if not isinstance(manifest_obj, dict):
        raise VeldraValidationError("manifest.json must deserialize to an object.")
    if not isinstance(config_obj, dict):
        raise VeldraValidationError("run_config.yaml must deserialize to an object.")
    if not isinstance(feature_schema_obj, dict):
        raise VeldraValidationError("feature_schema.json must deserialize to an object.")

    feature_names_raw = feature_schema_obj.get("feature_names")
    feature_names = (
        [str(item) for item in feature_names_raw]
        if isinstance(feature_names_raw, list)
        else []
    )
    feature_dtypes_raw = feature_schema_obj.get("feature_dtypes")
    feature_dtypes = (
        {str(k): str(v) for k, v in feature_dtypes_raw.items()}
        if isinstance(feature_dtypes_raw, dict)
        else {}
    )
    task_cfg = config_obj.get("task") if isinstance(config_obj.get("task"), dict) else {}
    data_cfg = config_obj.get("data") if isinstance(config_obj.get("data"), dict) else {}
    target_col = str(feature_schema_obj.get("target") or data_cfg.get("target") or "")
    task_type = str(
        manifest_obj.get("task_type")
        or task_cfg.get("type")
        or feature_schema_obj.get("task_type")
        or "unknown"
    )

    train_metrics: dict[str, float] = {}
    if metrics_path.exists():
        try:
            metrics_obj = json.loads(metrics_path.read_text(encoding="utf-8"))
            train_metrics = _flatten_numeric_metrics(metrics_obj)
        except Exception:
            train_metrics = {}

    return ArtifactSpec(
        artifact_path=str(path),
        run_id=str(manifest_obj.get("run_id", path.name)),
        task_type=task_type,
        target_col=target_col,
        feature_names=feature_names,
        feature_dtypes=feature_dtypes,
        train_metrics=train_metrics,
        created_at_utc=(
            str(manifest_obj["created_at_utc"])
            if manifest_obj.get("created_at_utc") is not None
            else None
        ),
    )


def validate_prediction_data(
    artifact_spec: ArtifactSpec,
    data_path: str,
) -> list[GuardRailResult]:
    inspected = inspect_data(data_path)
    if not inspected.get("success"):
        return [
            GuardRailResult(
                "error",
                f"Failed to inspect inference data: {inspected.get('error', 'unknown error')}",
            )
        ]

    stats = inspected.get("stats") or {}
    actual_cols = [str(col) for col in stats.get("columns", [])]
    actual_set = set(actual_cols)
    expected = [str(col) for col in artifact_spec.feature_names]
    expected_set = set(expected)

    findings: list[GuardRailResult] = []
    missing = sorted(expected_set - actual_set)
    extra = sorted(actual_set - expected_set)
    if missing:
        findings.append(
            GuardRailResult(
                "error",
                f"Missing required feature columns: {', '.join(missing[:10])}",
                "Align input columns with artifact feature_schema before prediction.",
            )
        )
    if extra:
        findings.append(
            GuardRailResult(
                "info",
                f"Extra columns detected (ignored during predict): {', '.join(extra[:10])}",
            )
        )

    profile_map: dict[str, str] = {}
    for profile in stats.get("column_profiles", []):
        if not isinstance(profile, dict):
            continue
        name = str(profile.get("name") or "")
        dtype_name = str(profile.get("dtype") or "")
        if name:
            profile_map[name] = dtype_name

    if artifact_spec.feature_dtypes:
        mismatch: list[str] = []
        for name in expected:
            expected_dtype = artifact_spec.feature_dtypes.get(name)
            actual_dtype = profile_map.get(name)
            if not expected_dtype or not actual_dtype:
                continue
            if _normalize_dtype_bucket(expected_dtype) != _normalize_dtype_bucket(actual_dtype):
                mismatch.append(f"{name} (expected={expected_dtype}, actual={actual_dtype})")
        if mismatch:
            findings.append(
                GuardRailResult(
                    "warning",
                    f"Potential dtype mismatches: {', '.join(mismatch[:8])}",
                    "Check feature engineering and column dtype casting before prediction.",
                )
            )
    else:
        findings.append(
            GuardRailResult(
                "info",
                "feature_schema.feature_dtypes is unavailable; strict dtype check is skipped.",
            )
        )

    if not findings:
        findings.append(GuardRailResult("ok", "Prediction data checks passed."))
    return findings


def delete_artifact_dir(
    artifact_path: str,
    *,
    root_dir: str = "artifacts",
) -> str:
    target = Path(_require(artifact_path, "artifact_path")).resolve()
    root = Path(_require(root_dir, "root_dir")).resolve()
    if not target.exists() or not target.is_dir():
        raise VeldraValidationError(f"Artifact directory does not exist: {target}")
    if target == root:
        raise VeldraValidationError("Refusing to delete artifact root directory.")
    if not target.is_relative_to(root):
        raise VeldraValidationError("Deletion is allowed only under artifacts root.")
    shutil.rmtree(target)
    return str(target)


@dataclass(slots=True)
class GuardRailResult:
    level: Literal["error", "warning", "info", "ok"]
    message: str
    suggestion: str | None = None


class GuardRailChecker:
    """Best-effort pre-run diagnostics for GUI pages."""

    def check_target(
        self,
        data: pd.DataFrame,
        target_col: str | None,
        task_type: str | None,
        *,
        exclude_cols: list[str] | None = None,
    ) -> list[GuardRailResult]:
        results: list[GuardRailResult] = []
        if not target_col:
            return [GuardRailResult("error", "Target column is required.")]
        if target_col not in data.columns:
            return [GuardRailResult("error", f"Target column not found: {target_col}")]

        exclude_cols = exclude_cols or []
        if target_col in exclude_cols:
            results.append(
                GuardRailResult(
                    "error",
                    "Target column is included in exclude columns.",
                    "Remove target from exclude list.",
                )
            )

        tgt = data[target_col]
        null_rate = float(tgt.isna().mean())
        if null_rate > 0.05:
            results.append(
                GuardRailResult(
                    "warning",
                    f"Target has missing values: {null_rate:.1%}.",
                    "Clean missing labels before training.",
                )
            )

        unique_count = int(tgt.dropna().nunique())
        if task_type == "binary" and unique_count != 2:
            results.append(
                GuardRailResult(
                    "error",
                    f"Binary task selected but unique target values = {unique_count}.",
                    "Use multiclass or adjust target encoding.",
                )
            )
        if task_type == "multiclass" and unique_count > 50:
            results.append(
                GuardRailResult(
                    "warning",
                    f"Multiclass with many classes ({unique_count}).",
                    "Consider class grouping or feature improvements.",
                )
            )

        if task_type in {"binary", "multiclass"}:
            freq = tgt.value_counts(normalize=True, dropna=True)
            if not freq.empty and float(freq.min()) < 0.05:
                results.append(
                    GuardRailResult(
                        "warning",
                        "Class imbalance detected (<5% minority class).",
                        "Enable auto class weight.",
                    )
                )

        if not results:
            results.append(GuardRailResult("ok", "Target checks passed."))
        return results

    def check_validation(
        self,
        data: pd.DataFrame,
        split_config: dict[str, Any],
        *,
        task_type: str | None = None,
        exclude_cols: list[str] | None = None,
    ) -> list[GuardRailResult]:
        results: list[GuardRailResult] = []
        split_type = str(split_config.get("type", "kfold"))
        n_splits = int(split_config.get("n_splits", 5))
        exclude_cols = exclude_cols or []

        if split_type == "timeseries" and not split_config.get("time_col"):
            results.append(
                GuardRailResult(
                    "error",
                    "Time Series split requires a time column.",
                    "Select a time column in Validation.",
                )
            )

        if split_type == "group" and not split_config.get("group_col"):
            results.append(
                GuardRailResult(
                    "error",
                    "Group split requires a group column.",
                    "Select a group column in Validation.",
                )
            )

        if n_splits > max(2, len(data) // 10):
            results.append(
                GuardRailResult(
                    "warning",
                    f"n_splits={n_splits} may be too high for {len(data)} rows.",
                    "Lower fold count or provide more data.",
                )
            )

        if split_type != "timeseries":
            for col in data.columns:
                if col in exclude_cols:
                    continue
                if pd.api.types.is_datetime64_any_dtype(data[col]):
                    results.append(
                        GuardRailResult(
                            "warning",
                            f"Datetime feature '{col}' may leak future information.",
                            "Use timeseries split or exclude datetime features.",
                        )
                    )
                    break

        if task_type in {"binary", "multiclass"} and split_type == "kfold":
            results.append(
                GuardRailResult(
                    "info",
                    "Stratified split is usually better for classification.",
                    "Switch split type to stratified when possible.",
                )
            )

        if not results:
            results.append(GuardRailResult("ok", "Validation checks passed."))
        return results

    def check_train(self, config: dict[str, Any]) -> list[GuardRailResult]:
        results: list[GuardRailResult] = []
        lr = float(config.get("learning_rate", 0.05))
        rounds = int(config.get("num_boost_round", 300))
        if lr > 0.3:
            results.append(
                GuardRailResult(
                    "warning",
                    f"Learning rate is high ({lr}).",
                    "Consider <= 0.1 for stable training.",
                )
            )
        if rounds > 5000:
            results.append(
                GuardRailResult(
                    "warning",
                    f"num_boost_round is large ({rounds}).",
                    "Verify early stopping settings.",
                )
            )
        if not results:
            results.append(GuardRailResult("ok", "Train checks passed."))
        return results

    def check_pre_run(self, config_yaml: str, data_path: str | None) -> list[GuardRailResult]:
        results: list[GuardRailResult] = []
        if not data_path or not Path(data_path).exists():
            results.append(
                GuardRailResult(
                    "error",
                    f"Data file not found: {data_path or 'None'}",
                    "Select a valid dataset path.",
                )
            )
        try:
            validate_config(config_yaml)
        except Exception as exc:
            results.append(GuardRailResult("error", f"Config validation error: {exc}"))
        if not results:
            results.append(GuardRailResult("ok", "Pre-run checks passed."))
        return results


def _flatten_numeric_metrics(metrics: Any) -> dict[str, float]:
    if not isinstance(metrics, dict):
        return {}
    top = {k: float(v) for k, v in metrics.items() if isinstance(v, (int, float))}
    if top:
        return top
    mean_obj = metrics.get("mean")
    if isinstance(mean_obj, dict):
        return {k: float(v) for k, v in mean_obj.items() if isinstance(v, (int, float))}
    return {}


def _export_output_path(artifact_path: str, suffix: str) -> Path:
    parent = Path(artifact_path).resolve()
    out_dir = parent / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(UTC).astimezone(_JST).strftime("%Y%m%d_%H%M%S")
    return out_dir / f"{suffix}_{ts}"


def export_excel_report(artifact_path: str) -> str:
    artifact = Artifact.load(_require(artifact_path, "artifact_path"))
    output_path = _export_output_path(artifact_path, "report").with_suffix(".xlsx")

    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise VeldraValidationError(f"Excel export requires openpyxl: {exc}")

    workbook = Workbook()
    sheet_metrics = workbook.active
    sheet_metrics.title = "metrics"
    sheet_metrics.append(["metric", "value"])
    metrics = _flatten_numeric_metrics(getattr(artifact, "metrics", None) or {})
    for key, value in sorted(metrics.items()):
        sheet_metrics.append([key, value])

    sheet_cfg = workbook.create_sheet("config")
    cfg_text = yaml.safe_dump(_to_safe_dict(getattr(artifact, "config", {})), sort_keys=False)
    for idx, line in enumerate(cfg_text.splitlines(), start=1):
        sheet_cfg.cell(row=idx, column=1, value=line)

    sheet_shap = workbook.create_sheet("shap")
    try:
        import shap  # noqa: F401

        sheet_shap.append(["status", "SHAP generation deferred (not computed in GUI MVP)."])
    except Exception:
        sheet_shap.append(["status", "SHAP dependency not installed; sheet skipped."])

    workbook.save(output_path)
    return str(output_path)


def export_html_report(artifact_path: str) -> str:
    artifact = Artifact.load(_require(artifact_path, "artifact_path"))
    output_path = _export_output_path(artifact_path, "report").with_suffix(".html")
    metrics = _flatten_numeric_metrics(getattr(artifact, "metrics", None) or {})
    cfg_text = yaml.safe_dump(_to_safe_dict(getattr(artifact, "config", {})), sort_keys=False)
    task_type = getattr(artifact, "task_type", "unknown")
    run_id = getattr(artifact, "run_id", "unknown")
    fold_metrics = getattr(artifact, "fold_metrics", None)
    causal_summary = load_causal_summary(artifact_path)

    rows = "".join(f"<tr><td>{k}</td><td>{v:.6g}</td></tr>" for k, v in sorted(metrics.items()))
    fold_rows = ""
    if isinstance(fold_metrics, pd.DataFrame) and not fold_metrics.empty:
        limited = fold_metrics.head(20)
        for _, row in limited.iterrows():
            cells = "".join(f"<td>{row[col]}</td>" for col in limited.columns)
            fold_rows += f"<tr>{cells}</tr>"

    causal_rows = ""
    if isinstance(causal_summary, dict):
        keys = [
            "method",
            "estimand",
            "estimate",
            "overlap_metric",
            "smd_max_unweighted",
            "smd_max_weighted",
        ]
        for key in keys:
            if key in causal_summary:
                causal_rows += f"<tr><td>{key}</td><td>{causal_summary[key]}</td></tr>"

    html_text = (
        "<html><head><meta charset='utf-8'><title>Veldra Report</title></head><body>"
        f"<h1>Veldra Report</h1><p>Run ID: {run_id} | Task: {task_type}</p>"
        "<h2>Metrics</h2><table border='1' cellpadding='4'>"
        "<tr><th>Metric</th><th>Value</th></tr>"
        f"{rows}</table>"
        "<h2>Fold Metrics (Preview)</h2><table border='1' cellpadding='4'>"
        f"{fold_rows or '<tr><td>Not available</td></tr>'}</table>"
        "<h2>Causal Diagnostics</h2><table border='1' cellpadding='4'>"
        f"{causal_rows or '<tr><td>Not available</td></tr>'}</table>"
        "<h2>Config</h2><pre>"
        f"{cfg_text}"
        "</pre>"
        "</body></html>"
    )
    try:
        from jinja2 import Template

        tmpl = Template(html_text)
        html_text = tmpl.render()
    except Exception:
        pass
    output_path.write_text(html_text, encoding="utf-8")
    return str(output_path)


def export_pdf_report(artifact_path: str) -> str:
    html_path = export_html_report(artifact_path)
    pdf_path = Path(html_path).with_suffix(".pdf")
    try:
        from weasyprint import HTML  # type: ignore
    except Exception as exc:
        raise VeldraValidationError(
            "PDF export requires optional dependency 'weasyprint'. "
            "Install it to enable export_pdf_report."
        ) from exc
    HTML(filename=str(html_path)).write_pdf(str(pdf_path))
    return str(pdf_path)


def _to_safe_dict(value: Any) -> dict[str, Any]:
    if is_dataclass(value):
        return asdict(value)
    if hasattr(value, "model_dump"):
        try:
            dumped = value.model_dump(mode="json")
            if isinstance(dumped, dict):
                return dumped
        except Exception:
            pass
    if isinstance(value, dict):
        return value
    return {"value": repr(value)}


@dataclass(slots=True)
class _RunActionContext:
    job_id: str | None = None
    job_store: GuiJobStore | None = None
    action: str | None = None
    tune_total_trials: int | None = None

    def update_progress(self, pct: float, step: str) -> None:
        if not self.job_id or self.job_store is None:
            return
        self.job_store.update_progress(self.job_id, pct, step=step)

    def append_log(
        self,
        *,
        level: str,
        message: str,
        payload: dict[str, Any] | None = None,
    ) -> None:
        if not self.job_id or self.job_store is None:
            return
        self.job_store.append_job_log(
            self.job_id,
            level=level,
            message=message,
            payload=payload,
        )

    def on_runner_payload(self, level: str, message: str, payload: dict[str, Any]) -> None:
        self.append_log(level=level, message=message, payload=payload)
        if (
            self.action == "tune"
            and self.tune_total_trials is not None
            and self.tune_total_trials > 0
            and "n_trials_done" in payload
        ):
            done = int(payload.get("n_trials_done", 0))
            total = int(self.tune_total_trials)
            pct = 20.0 + (70.0 * max(0, min(done, total)) / float(total))
            self.update_progress(pct, f"tuning_trials_{done}/{total}")

    def check_cancellation(self, step: str = "checkpoint") -> None:
        if not self.job_id or self.job_store is None:
            return
        if self.job_store.is_cancel_requested(self.job_id):
            self.update_progress(100.0, "canceled")
            self.append_log(
                level="WARNING",
                message="action_canceled",
                payload={"step": step},
            )
            raise CanceledByUser("Canceled by user request.")

    def cancellation_hook(self) -> None:
        self.check_cancellation("runner_hook")


class _RunnerLogCapture(logging.Handler):
    def __init__(self, context: _RunActionContext) -> None:
        super().__init__()
        self._context = context

    def emit(self, record: logging.LogRecord) -> None:
        payload: dict[str, Any] = {}
        raw = record.getMessage()
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                payload = parsed
        except Exception:
            payload = {"raw": raw}
        message = str(getattr(record, "event_message", "runner_event"))
        self._context.on_runner_payload(record.levelname, message, payload)


@contextmanager
def _capture_runner_logs(context: _RunActionContext):
    logger = logging.getLogger("veldra")
    handler = _RunnerLogCapture(context)
    logger.addHandler(handler)
    try:
        yield
    finally:
        logger.removeHandler(handler)


def run_action(
    invocation: RunInvocation,
    *,
    job_id: str | None = None,
    job_store: GuiJobStore | None = None,
) -> GuiRunResult:
    try:
        action = invocation.action.strip().lower()
        context = _RunActionContext(
            job_id=job_id,
            job_store=job_store,
            action=action,
        )
        context.update_progress(5.0, "validating")
        context.append_log(level="INFO", message="action_started", payload={"action": action})
        context.check_cancellation("validated")
        if action not in {
            "fit",
            "evaluate",
            "predict",
            "tune",
            "simulate",
            "export",
            "estimate_dr",
            "export_excel",
            "export_html_report",
            "export_pdf_report",
        }:
            raise VeldraValidationError(f"Unsupported action '{invocation.action}'.")

        if action == "fit":
            context.update_progress(12.0, "load_config")
            config = _resolve_config(invocation)
            context.check_cancellation("fit_after_config")
            context.update_progress(20.0, "fit_running")
            with _capture_runner_logs(context):
                result = _invoke_runner_with_optional_hook(
                    _get_runner_func("fit"),
                    config,
                    cancellation_hook=context.cancellation_hook,
                )
            context.check_cancellation("fit_after_runner")
            context.update_progress(95.0, "fit_finalize")
        elif action == "tune":
            context.update_progress(12.0, "load_config")
            config = _resolve_config(invocation)
            context.tune_total_trials = int(config.tuning.n_trials)
            context.check_cancellation("tune_after_config")
            context.update_progress(20.0, "tune_running")
            with _capture_runner_logs(context):
                result = _invoke_runner_with_optional_hook(
                    _get_runner_func("tune"),
                    config,
                    cancellation_hook=context.cancellation_hook,
                )
            context.check_cancellation("tune_after_runner")
            context.update_progress(95.0, "tune_finalize")
        elif action == "estimate_dr":
            context.update_progress(12.0, "load_config")
            config = _resolve_config(invocation)
            context.check_cancellation("estimate_after_config")
            context.update_progress(20.0, "estimate_running")
            with _capture_runner_logs(context):
                result = _invoke_runner_with_optional_hook(
                    _get_runner_func("estimate_dr"),
                    config,
                    cancellation_hook=context.cancellation_hook,
                )
            context.check_cancellation("estimate_after_runner")
            context.update_progress(95.0, "estimate_finalize")
        elif action == "evaluate":
            context.update_progress(12.0, "load_data")
            data_path = _require(invocation.data_path, "data_path")
            frame = _get_load_tabular_data()(data_path)
            context.check_cancellation("evaluate_after_data")
            if invocation.artifact_path and invocation.artifact_path.strip():
                context.update_progress(30.0, "load_artifact")
                artifact = Artifact.load(invocation.artifact_path.strip())
                context.check_cancellation("evaluate_after_artifact")
                context.update_progress(45.0, "evaluate_running")
                with _capture_runner_logs(context):
                    result = _get_runner_func("evaluate")(artifact, frame)
            else:
                context.update_progress(30.0, "load_config")
                config = _resolve_config(invocation)
                context.check_cancellation("evaluate_after_config")
                context.update_progress(45.0, "evaluate_running")
                with _capture_runner_logs(context):
                    result = _get_runner_func("evaluate")(config, frame)
            context.check_cancellation("evaluate_after_runner")
            context.update_progress(95.0, "evaluate_finalize")
        elif action == "predict":
            context.update_progress(12.0, "load_data")
            data_path = _require(invocation.data_path, "data_path")
            frame = _get_load_tabular_data()(data_path)
            context.check_cancellation("predict_after_data")
            context.update_progress(30.0, "load_artifact")
            artifact_path = _require(invocation.artifact_path, "artifact_path")
            artifact = Artifact.load(artifact_path)
            context.check_cancellation("predict_after_artifact")
            context.update_progress(45.0, "predict_running")
            with _capture_runner_logs(context):
                prediction = _get_runner_func("predict")(artifact, frame)

            pred_data = getattr(prediction, "data", None)
            if isinstance(pred_data, pd.DataFrame):
                pred_df = pred_data.reset_index(drop=True)
            else:
                pred_series = pd.Series(pred_data, name="prediction")
                pred_df = pred_series.to_frame().reset_index(drop=True)

            suffix = job_id or str(int(datetime.now(UTC).timestamp() * 1_000_000_000))
            out_path = _studio_predict_tmp_dir() / f"predict_{suffix}.csv"
            pred_df.to_csv(out_path, index=False)
            result = {
                "task_type": str(getattr(prediction, "task_type", "")),
                "metadata": (
                    getattr(prediction, "metadata", {})
                    if isinstance(getattr(prediction, "metadata", {}), dict)
                    else {}
                ),
                "prediction_csv_path": str(out_path),
                "preview_rows": pred_df.head(100).to_dict(orient="records"),
                "total_count": int(len(pred_df)),
            }
            context.check_cancellation("predict_after_runner")
            context.update_progress(95.0, "predict_finalize")
        elif action == "simulate":
            context.update_progress(12.0, "load_inputs")
            artifact_path = _require(invocation.artifact_path, "artifact_path")
            data_path = _require(invocation.data_path, "data_path")
            scenarios_path = _require(invocation.scenarios_path, "scenarios_path")
            artifact = Artifact.load(artifact_path)
            frame = _get_load_tabular_data()(data_path)
            scenarios = _load_scenarios(scenarios_path)
            context.check_cancellation("simulate_after_inputs")
            context.update_progress(45.0, "simulate_running")
            with _capture_runner_logs(context):
                result = _get_runner_func("simulate")(artifact, frame, scenarios)
            context.check_cancellation("simulate_after_runner")
            context.update_progress(95.0, "simulate_finalize")
        elif action == "export_excel":
            context.update_progress(15.0, "load_artifact")
            artifact_path = _require(invocation.artifact_path, "artifact_path")
            context.check_cancellation("export_excel_after_artifact")
            context.update_progress(55.0, "export_excel_running")
            output_path = export_excel_report(artifact_path)
            result = {"artifact_path": artifact_path, "output_path": output_path}
            context.check_cancellation("export_excel_after_runner")
            context.update_progress(95.0, "export_excel_finalize")
        elif action == "export_html_report":
            context.update_progress(15.0, "load_artifact")
            artifact_path = _require(invocation.artifact_path, "artifact_path")
            context.check_cancellation("export_html_after_artifact")
            context.update_progress(55.0, "export_html_running")
            output_path = export_html_report(artifact_path)
            result = {"artifact_path": artifact_path, "output_path": output_path}
            context.check_cancellation("export_html_after_runner")
            context.update_progress(95.0, "export_html_finalize")
        elif action == "export_pdf_report":
            context.update_progress(15.0, "load_artifact")
            artifact_path = _require(invocation.artifact_path, "artifact_path")
            context.check_cancellation("export_pdf_after_artifact")
            context.update_progress(55.0, "export_pdf_running")
            output_path = export_pdf_report(artifact_path)
            result = {"artifact_path": artifact_path, "output_path": output_path}
            context.check_cancellation("export_pdf_after_runner")
            context.update_progress(95.0, "export_pdf_finalize")
        else:
            context.update_progress(15.0, "load_artifact")
            artifact_path = _require(invocation.artifact_path, "artifact_path")
            artifact = Artifact.load(artifact_path)
            export_format = (invocation.export_format or "python").strip().lower()
            context.check_cancellation("export_after_artifact")
            context.update_progress(55.0, "export_running")
            with _capture_runner_logs(context):
                result = _get_runner_func("export")(artifact, format=export_format)
            context.check_cancellation("export_after_runner")
            context.update_progress(95.0, "export_finalize")

        context.check_cancellation("before_completion")
        context.update_progress(100.0, "completed")
        context.append_log(level="INFO", message="action_completed", payload={"action": action})
        return GuiRunResult(
            success=True,
            message=f"{action} completed successfully.",
            payload=_result_to_payload(result),
        )
    except CanceledByUser:
        raise
    except Exception as exc:
        if job_id and job_store is not None:
            context = _RunActionContext(
                job_id=job_id,
                job_store=job_store,
                action=invocation.action,
            )
            payload = _build_error_payload(exc, action=invocation.action)
            context.append_log(
                level="ERROR",
                message="action_failed",
                payload={"error": normalize_gui_error(exc), **payload},
            )
        payload = _build_error_payload(exc, action=invocation.action)
        return GuiRunResult(success=False, message=normalize_gui_error(exc), payload=payload)


def submit_run_job(invocation: RunInvocation) -> GuiJobResult:
    store = get_gui_job_store()
    job = store.enqueue_job(invocation)
    _start_worker_if_needed()
    log_event(
        LOGGER,
        logging.INFO,
        "gui job submitted",
        run_id=job.job_id,
        artifact_path=invocation.artifact_path,
        task_type=job.action,
        event="gui job submitted",
        status=job.status,
        action=job.action,
    )
    return GuiJobResult(
        job_id=job.job_id,
        status=job.status,
        message=f"Queued {job.action} job.",
    )


def get_run_job(job_id: str) -> GuiJobRecord | None:
    return get_gui_job_store().get_job(_require(job_id, "job_id"))


def list_run_jobs(*, limit: int = 100, status: str | None = None) -> list[GuiJobRecord]:
    rows, _total = get_gui_job_store().list_jobs_page(
        limit=limit,
        offset=0,
        status=status,
    )
    return rows


def list_run_jobs_page(
    *,
    limit: int = 50,
    offset: int = 0,
    status: str | None = None,
    action: str | None = None,
    query: str | None = None,
) -> PaginatedResult[GuiJobRecord]:
    rows, total_count = get_gui_job_store().list_jobs_page(
        limit=limit,
        offset=offset,
        status=status,
        action=action,
        query=query,
    )
    return PaginatedResult[GuiJobRecord](
        items=rows,
        total_count=total_count,
        limit=max(1, int(limit)),
        offset=max(0, int(offset)),
    )


def list_run_job_logs(
    job_id: str,
    *,
    limit: int = 200,
    after_seq: int | None = None,
) -> list[dict[str, Any]]:
    records = get_gui_job_store().list_job_logs(
        _require(job_id, "job_id"),
        limit=limit,
        after_seq=after_seq,
    )
    return [
        {
            "job_id": row.job_id,
            "seq": row.seq,
            "created_at_utc": row.created_at_utc,
            "level": row.level,
            "message": row.message,
            "payload": row.payload,
        }
        for row in records
    ]


def cancel_run_job(job_id: str) -> GuiJobResult:
    canceled = get_gui_job_store().request_cancel(_require(job_id, "job_id"))
    if canceled is None:
        raise VeldraValidationError(f"Job not found: {job_id}")
    log_event(
        LOGGER,
        logging.INFO,
        "gui job updated",
        run_id=canceled.job_id,
        artifact_path=canceled.invocation.artifact_path,
        task_type=canceled.action,
        event="gui job updated",
        status=canceled.status,
        action=canceled.action,
        cancel_requested=canceled.cancel_requested,
    )
    return GuiJobResult(
        job_id=canceled.job_id,
        status=canceled.status,
        message=f"Job {canceled.job_id} status updated to {canceled.status}.",
    )


def retry_run_job(job_id: str) -> GuiJobResult:
    source = get_gui_job_store().get_job(_require(job_id, "job_id"))
    if source is None:
        raise VeldraValidationError(f"Job not found: {job_id}")
    if source.status not in {"failed", "canceled"}:
        raise VeldraValidationError(
            f"Retry is allowed only for failed/canceled jobs (status={source.status})."
        )
    policy = source.invocation.retry_policy or RetryPolicy()
    retried = get_gui_job_store().create_retry_job(
        source.job_id,
        reason="manual_retry",
        policy=policy,
    )
    log_event(
        LOGGER,
        logging.INFO,
        "gui job retried",
        run_id=retried.job_id,
        artifact_path=retried.invocation.artifact_path,
        task_type=retried.action,
        event="gui job retried",
        status=retried.status,
        action=retried.action,
        source_job_id=source.job_id,
        retry_count=retried.retry_count,
    )
    return GuiJobResult(
        job_id=retried.job_id,
        status=retried.status,
        message=f"Retry queued from {source.job_id} as {retried.job_id}.",
    )


def set_run_job_priority(job_id: str, priority: str | GuiJobPriority) -> GuiJobResult:
    normalized = str(priority).strip().lower()
    if normalized not in {"low", "normal", "high"}:
        raise VeldraValidationError(f"Unsupported priority: {priority}")
    updated = get_gui_job_store().set_job_priority(
        _require(job_id, "job_id"),
        normalized,
    )
    if updated is None:
        raise VeldraValidationError(f"Job not found: {job_id}")
    log_event(
        LOGGER,
        logging.INFO,
        "gui job priority updated",
        run_id=updated.job_id,
        artifact_path=updated.invocation.artifact_path,
        task_type=updated.action,
        event="gui job priority updated",
        status=updated.status,
        action=updated.action,
        priority=updated.priority,
    )
    return GuiJobResult(
        job_id=updated.job_id,
        status=updated.status,
        message=f"Job {updated.job_id} priority updated to {updated.priority}.",
    )


def list_artifacts(root_dir: str) -> list[ArtifactSummary]:
    return list_artifacts_page(root_dir=root_dir, limit=10_000, offset=0).items


def list_artifacts_page(
    *,
    root_dir: str,
    limit: int = 50,
    offset: int = 0,
    query: str | None = None,
) -> PaginatedResult[ArtifactSummary]:
    started_at = perf_counter()
    root = Path(_require(root_dir, "root_dir"))
    if not root.exists():
        return PaginatedResult[ArtifactSummary](
            items=[],
            total_count=0,
            limit=max(1, int(limit)),
            offset=max(0, int(offset)),
        )
    if not root.is_dir():
        raise VeldraValidationError(f"Artifact root is not a directory: {root}")

    summaries: list[ArtifactSummary] = []
    q = str(query or "").strip().lower()
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        manifest_path = child / "manifest.json"
        if not manifest_path.exists():
            continue
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            item = ArtifactSummary(
                path=str(child),
                run_id=str(manifest.get("run_id", child.name)),
                task_type=str(manifest.get("task_type", "unknown")),
                created_at_utc=(
                    str(manifest["created_at_utc"])
                    if manifest.get("created_at_utc") is not None
                    else None
                ),
            )
        except Exception:
            item = ArtifactSummary(
                path=str(child),
                run_id=child.name,
                task_type="unknown",
                created_at_utc=None,
            )
        if (
            not q
            or q in item.run_id.lower()
            or q in item.task_type.lower()
            or q in item.path.lower()
        ):
            summaries.append(item)

    summaries.sort(key=lambda item: item.created_at_utc or "", reverse=True)
    safe_limit = max(1, min(int(limit), 1000))
    safe_offset = max(0, int(offset))
    paged = summaries[safe_offset : safe_offset + safe_limit]
    result = PaginatedResult[ArtifactSummary](
        items=paged,
        total_count=len(summaries),
        limit=safe_limit,
        offset=safe_offset,
    )
    record_perf_metric(
        "list_artifacts_page",
        (perf_counter() - started_at) * 1000.0,
        {
            "root_dir": str(root),
            "offset": safe_offset,
            "limit": safe_limit,
            "total_count": result.total_count,
            "query": bool(q),
        },
    )
    return result


def list_run_jobs_filtered(
    *,
    limit: int = 200,
    status: str | None = None,
    action: str | None = None,
    query: str | None = None,
) -> list[GuiJobRecord]:
    page = list_run_jobs_page(
        limit=limit,
        offset=0,
        status=status,
        action=action,
        query=query,
    )
    filtered = list(page.items)
    if action:
        action_norm = str(action).strip().lower()
        filtered = [job for job in filtered if job.action == action_norm]
    if query:
        q = str(query).strip().lower()
        filtered = [
            job
            for job in filtered
            if q in job.job_id.lower()
            or q in (job.invocation.artifact_path or "").lower()
            or q in (job.invocation.config_path or "").lower()
        ]
    return filtered


def delete_run_jobs(job_ids: list[str]) -> int:
    return get_gui_job_store().delete_jobs(job_ids)


def load_job_config_yaml(job: GuiJobRecord) -> str:
    if job.invocation.config_yaml and job.invocation.config_yaml.strip():
        return job.invocation.config_yaml
    if job.invocation.config_path and job.invocation.config_path.strip():
        return load_config_yaml(job.invocation.config_path)
    raise VeldraValidationError(f"No config source available for job: {job.job_id}")


def compare_artifacts(artifact_a: str, artifact_b: str) -> dict[str, Any]:
    return compare_artifacts_multi([artifact_a, artifact_b], baseline=artifact_a)


def load_causal_summary(artifact_path: str) -> dict[str, Any] | None:
    base = Path(_require(artifact_path, "artifact_path"))
    candidates: list[Path] = []
    direct = base / "dr_summary.json"
    if direct.exists():
        candidates.append(direct)
    parent = base.parent
    causal_root = parent / "causal"
    if causal_root.exists() and causal_root.is_dir():
        for summary in causal_root.glob("*/dr_summary.json"):
            candidates.append(summary)
    if not candidates:
        return None
    latest = max(candidates, key=lambda p: p.stat().st_mtime)
    try:
        payload = json.loads(latest.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def compare_artifacts_multi(artifacts: list[str], baseline: str) -> dict[str, Any]:
    selected = [str(p) for p in artifacts if str(p).strip()]
    if not selected:
        raise VeldraValidationError("At least one artifact is required.")
    if len(selected) > 5:
        raise VeldraValidationError("At most 5 artifacts can be compared.")
    if baseline not in selected:
        baseline = selected[0]

    loaded = {path: Artifact.load(path) for path in selected}
    checks: list[dict[str, str]] = []
    task_types = {
        path: str(
            getattr(
                art,
                "task_type",
                getattr(getattr(art, "manifest", None), "task_type", "unknown"),
            )
        )
        for path, art in loaded.items()
    }
    if len(set(task_types.values())) == 1:
        checks.append(
            {
                "level": "ok",
                "message": f"Same task type: {next(iter(task_types.values()))}",
            }
        )
    else:
        checks.append({"level": "warning", "message": "Different task types detected."})

    cfg_yamls: dict[str, str] = {}
    cfg_payloads: dict[str, dict[str, Any]] = {}
    metrics_map: dict[str, dict[str, float]] = {}
    for path, art in loaded.items():
        cfg_obj = _to_safe_dict(getattr(art, "config", {}) or getattr(art, "run_config", {}))
        cfg_payloads[path] = cfg_obj if isinstance(cfg_obj, dict) else {}
        cfg_yamls[path] = yaml.safe_dump(cfg_obj, sort_keys=False)
        metrics_map[path] = _flatten_numeric_metrics(getattr(art, "metrics", {}) or {})

    baseline_cfg = cfg_payloads.get(baseline, {})
    baseline_data_obj = (baseline_cfg.get("data") or {}) if isinstance(baseline_cfg, dict) else {}
    baseline_data = baseline_data_obj.get("path")
    baseline_split = (
        (baseline_cfg.get("split") or {}) if isinstance(baseline_cfg, dict) else {}
    ).get("type")
    for path in selected:
        if path == baseline:
            continue
        cfg_obj = cfg_payloads.get(path, {})
        data_path = ((cfg_obj.get("data") or {}) if isinstance(cfg_obj, dict) else {}).get("path")
        split_type = ((cfg_obj.get("split") or {}) if isinstance(cfg_obj, dict) else {}).get("type")
        if baseline_data and data_path and baseline_data != data_path:
            checks.append(
                {
                    "level": "warning",
                    "message": "Different data sources. Row-level prediction diff is disabled.",
                }
            )
        if baseline_split and split_type and baseline_split != split_type:
            checks.append(
                {
                    "level": "info",
                    "message": f"Split differs: {baseline_split} vs {split_type}",
                }
            )

    all_metrics = sorted({k for m in metrics_map.values() for k in m.keys()})
    baseline_metrics = metrics_map.get(baseline, {})
    metric_rows: list[dict[str, Any]] = []
    for metric in all_metrics:
        base_val = baseline_metrics.get(metric)
        for path in selected:
            cur_val = metrics_map[path].get(metric)
            delta = None
            if isinstance(cur_val, (int, float)) and isinstance(base_val, (int, float)):
                delta = float(cur_val) - float(base_val)
            metric_rows.append(
                {
                    "metric": metric,
                    "artifact": path,
                    "value": cur_val,
                    "baseline": base_val,
                    "delta_from_baseline": delta,
                }
            )

    first = selected[0]
    return {
        "checks": checks,
        "baseline": baseline,
        "artifacts": selected,
        "metrics_map": metrics_map,
        "metric_rows": metric_rows,
        "config_yaml_a": cfg_yamls.get(first, ""),
        "config_yaml_b": cfg_yamls.get(baseline, ""),
        "config_yamls": cfg_yamls,
    }
